#!/opt/anaconda3/bin/python

# dh-validate.py
#  
# A script to validate DataHarmonizer data files according to a given schema
# and optionally Class, in the case of data files that don't identify which 
# class their fields are from.  Passes data file to linkml-validate.  For this 
# reason, the script only works with LinkML installed.
#
# To prepare tsv/csv/xls/xlsx files for above validation, this script will 
# remove header lines until it encounters a line with every cell containing
# text, which is assumed to be the field name header.  (IN FUTURE this will 
# be sniffed better).  This includes removing section header line if any.
# Then a temporary file is created in yaml format, which adjusts number fields
# to be unquoted if linkml spec says its a number, and it is a valid number.
# As well, multivalued fields are split up into an array of separate values.
# cli linkml-validate is then applied to this temporary file.
#
# DataHarmonizer-generated data file with its section headers must be removed,
#
# 
# Options:
#  -s, --schema FILE               Schema file to validate data against
#  -C, --target-class TEXT 
#  -S, --index-slot TEXT    top level slot. Required for CSV dumping/loading
#  -V, --version                   Show the version and exit.
#
# > cd web/templates/[template folder]/
# > python ../../../script/dh-validate.py 
#
# Allowed data file types: '.tsv','.csv','.xls','.xlsx','.json','.yml','.yaml'
# Some formats (.json, .yml, .yaml) are evolving on DH side so not yet tested.
#
# dh-validate.py --schema schema.yaml test_data.csv
# dh-validate.py --schema schema.yaml --target-class "CanCOGeN Covid-19" some_data_file.csv
#
# Author: Damion Dooley
# 

from collections import OrderedDict
from decimal import Decimal
import csv
import re
from openpyxl import load_workbook # For .xlsx
import xlrd # for .xls // pip install xlrd
import yaml
import json
import pathlib
import os
from sys import exit
import argparse
from linkml_runtime.utils.schemaview import SchemaView
from linkml_runtime.dumpers.yaml_dumper import YAMLDumper
import subprocess

BOOLEANS = {'y':1, 'yes':1, 't':1, 'true':1, 'on':1, '1':1,
 'n':0, 'no':0, 'f':0, 'false':0, 'off':0, '0':0}

def init_parser():

	parser = argparse.ArgumentParser(
	  prog='dh-validate',
	  description='A wrapper around linkml-validate for validating data files against a LinkML schema.',
	  epilog='For more information, see https://github.com/cidgoh/DataHarmonizer/')

	# FUTURE: In case of json, schema will be provided in data file url, and 
	# potentially can be looked up dynamically.
	parser.add_argument("-s", "--schema", 
		dest="schema_path", 
		metavar="FILE",
		default= os.getcwd() +"/schema.yaml",
	  help='A LinkML yaml schema to test given data file against.  Default is "schema.yaml"');

	parser.add_argument("-C", "--target-class", 
		dest="target_class", 
		metavar="TEXT",
		required=False, 
	  help="A schema class to test given data file against.");

	parser.add_argument('data_sources', 
		metavar='DATA_SOURCES', 
		#nargs='+',
	  help="One or more data files to validate using given schema.");

	""" parser.add_argument("-V", "--version",
		dest="version",
		default=False,
		help="Return version # and then exit.",
	)
	"""

	return parser.parse_args();


# For a given schema, return the class that possibly fits data to be 
# validated. Also return the "index_slot" or likely identifier or 
# primary key slot, useful in converting tabular tsv / csv data.
#
# Currently tries to match inputted target_class if any.
# Alternately returns class if there is only one in schema.
# Future: could sniff data to see what matches best in case where 
# there are multiple classes.

def getTargetClass(SCHEMA, target_class, slot_key = None):
	templates = {};
	for name, class_obj in SCHEMA.all_classes().items():
	# Generate schema's list of possible template classes:
		if name == 'dh_interface': # Archaic class
			continue;
		templates[name] = class_obj;

	if target_class and not target_class in templates:
		exit("The given validation target class [" + target_class + "] was not found in schema! Schema has: " + str([key for key in templates.keys()]) );

	# So far many DH schemas only have one class so only possibility is to validate against that:
	if not target_class:
		if len(templates) == 1:
			target_class = next(iter(templates.keys())); 
		else:
			# FUTURE: determine if one of the multiple classes is a good match for given data.
			exit("No validation class was provided, and no default could be found!");

	# Determined target_class at this point.
	found = False;
	template = SCHEMA.get_class(target_class);
	for name in SCHEMA.class_slots(target_class):
		slot = template['attributes'][name];
		if 'identifier' in slot and slot['identifier'] == True:
			if slot_key:
				if slot_key != name:
					continue; # Continue search
				else:
					found = True;
					break;
			# Pick first eligable slot key if none provided on command line
			else:
				found = True;
				slot_key = name;

			break;

	print ("Identifier slot:", slot_key);
	return (target_class, slot_key);


# A dictionary of all the sections that may appear on first line of a DH data file
def getSlotGroupTitleDict(SCHEMA, target_class):
	
	slot_group_titles = {};

	template = SCHEMA.get_class(target_class);
	for name in SCHEMA.class_slots(target_class):
		slot = template['attributes'][name];
		#print("SLOT", slot);
		if slot['slot_group'] and slot['slot_group'] > '':
			slot_group_titles[slot['slot_group']] = True;

	return slot_group_titles;


def getSlotTitleToNameDict(SCHEMA, target_class):
	slot_title_name_map = {};

	template = SCHEMA.get_class(target_class);
	for name in SCHEMA.class_slots(target_class):
		slot = template['attributes'][name];
		slot_title_name_map[slot['title']] = name;

	return slot_title_name_map;


# Used to compose columns of new output file
# uses class.slots array to determine order, which should be same as 
# order in slot_usage and its rank.  FUTURE: ensure sort by slot_usage rank,
# if any.
def getSlotNameToTitleDict(SCHEMA, target_class):
	slot_name_title_map = OrderedDict();

	template = SCHEMA.get_class(target_class);
	for name in SCHEMA.class_slots(target_class):
		slot = template['attributes'][name];
		slot_name_title_map[name] = slot['title'] or None;

	return slot_name_title_map;


# LinkML validation doesn't care about order of fields, so we can add 
# missing schema headers in appropriate columns (according to ordering)
# with empty values. This enables a "new" tabular data file to be saved.
# subsequent data columns have to be mapped over however.

# Returns converted headers along with a report of mismatched headers which
# arise when old schema applied to newer data file or visa versa.
# In contrast to JSON, schema-version-appropriate tabular data should mention
# All columns/slots in SCHEMA
def getNormalizedHeaders(SCHEMA, target_class, row, slot_header_map, slot_title_map):

	report = OrderedDict();
	header_count = 0;

	# Data_row_map will always have a place for every place in row.
	data_row_map = [];
	ignored = OrderedDict();

	for field in row:
		if field in slot_header_map: # row already mentions slot name
			data_row_map.append(field);
			header_count +=1;
			continue;

		if field in slot_title_map: # row mentions slot title, so translate.
			data_row_map.append(slot_title_map[field]);
			report[field] = 'Mapped "' + field + '" to ' + slot_title_map[field];
			header_count +=1;	
			continue;

		# Possibly old naming that doesn't match title but would match name, try: 
		# try_name = Lower(regexreplace(regexreplace(field,"[ /]","_"),"[-()]","")).
			# ...

		else:
			data_row_map.append('');
			ignored[field] = 'Ignored "' + field + '", not in schema.';

	# TO DO: Determine how to handle rows with fewer or more unmatched columns.
	if header_count < len(row):
		exit ("Data file is missing fields: " + str(report));
		# FUTURE: allow shorter (empty tail) rows?  NO. 
		# take out unused columns?

	report |= ignored; # Add ignored fields to end of report.

	return (tuple(data_row_map), report, header_count);


# A normalized data file has tsv/csv/xls/xlsx files converted to JSON format.
# (This does not handle .yaml, .yml, .json-ld since those are validated ok
# with data elements saved by LinkML slot name rather than title.)
#
# FUTURE: handle situation where for every section there is only one slot name/title
def getNormalizedDataFile(SCHEMA, target_class, data_source, temp_base):

	if not os.path.isfile(data_source):
		exit("ERROR: Data file not found: " + data_source);

	# slot_group_titles = getSlotGroupTitleDict(SCHEMA, target_class);
	slot_title_map = getSlotTitleToNameDict(SCHEMA, target_class);
	slot_header_map = getSlotNameToTitleDict(SCHEMA, target_class);
	file_path_obj = pathlib.Path(data_source);
	target_class_CC = re.sub("[_ (/)-]","", target_class)# CamelCase version.
	reader = None;

	if file_path_obj.suffix in ['.xls','.xlsx']:
		file_mode = "rb"
	else:
		file_mode = "r"

	with open(data_source, file_mode) as data_handle:
		
		match file_path_obj.suffix:

			case '.yaml' | '.yml':
				exit("ERROR: getNormalizedDataFile() does not process yaml/yml files.");  

			# For CSV and TSV, if reader.fieldnames has empty labels, it means its
			# not a header line, i.e. skipping "section" line or other stuff.
			# (Don't even need to use "next(data_handle)")
			# Here we find reader.fieldnames, AND move reader TO FIRST ROW OF DATA.
			case '.csv' | '.tsv':
				while True:
					reader = csv.DictReader(data_handle, dialect = ('excel' if file_path_obj.suffix == ".csv" else 'excel-tab'));
					# Empty fieldname indicates superfluous initial row.
					if not '' in reader.fieldnames: 
						# Found row of full table cells.  Additional test on cell text = slot name/title?
						(header_row, report, header_count) = getNormalizedHeaders(SCHEMA, target_class, reader.fieldnames, slot_header_map, slot_title_map);
						break;

			# Excell types can hold multiple tabs, each of which needs to be validated separately.
			# Excel returns a tuple, one value position for each column.
			case '.xlsx':

				workbook = load_workbook(data_source);
				if target_class in workbook.sheetnames or target_class_CC in workbook.sheetnames: # e.g. ['Sheet1'], also note "sheet.title"
					sheet = workbook.worksheets[workbook.sheetnames.index(target_class_CC)]; 
				else:
					if len(workbook.sheetnames) > 1:
						exit("ERROR: getNormalizedDataFile() cannot find " + target_class + " tab in excel spreadsheet tabs: " + str(workbook.sheetnames)); 
					else:
						sheet = workbook.worksheets[0]; # pick the one and only tab/sheet.

				reader = sheet.iter_rows(values_only=True);

				for row in reader:
					# Skip superfluous header liness:
					if not '' in row:
						# Found row with value in each cell, so likely header.
						(header_row, report, header_count) = getNormalizedHeaders(SCHEMA, target_class, row, slot_header_map, slot_title_map);
						break;

			# Aligning .xls with same output as .xlsx
			case '.xls': 

				workbook = xlrd.open_workbook(data_source);
				sheets = workbook.nsheets;
				sheetnames = workbook.sheet_names();
				if target_class in sheetnames:
					sheet = workbook.sheet_by_index(sheetnames.index(target_class));
				else:
					if sheets > 1:
						exit("ERROR: getNormalizedDataFile() cannot find " + target_class + " tab in excel spreadsheet tabs: " + str(sheetnames)); 
					else:
						sheet = workbook.sheet_by_index(0);

				# Returns array of values for given row
				reader = iter(tuple([sheet.cell_value(rx, cx) for cx in range(sheet.ncols)]) for rx in range(sheet.nrows)); 
   
				for row in reader:
					# Skip superfluous header liness:
					if not '' in row:	
						(header_row, report, header_count) = getNormalizedHeaders(SCHEMA, target_class, row, slot_header_map, slot_title_map);
						break;

			case _:
				exit("ERROR: Data file doesn't have compatible type ('tsv/csv/xls/xlsx/json/yml/yaml) :" + data_source);

		# Write both normalized TSV and JSON files:

		yaml = writeTmpFiles (SCHEMA, target_class, header_row, reader, temp_base);

	return (yaml, report);


def writeTmpFiles (SCHEMA, target_class, header_row, reader, temp_base):
  # First row of DH tabular data may be slot_groups
  # 2nd DH row likely has column/field/slot names which need adjustment
  # (newline = '' prevents extra blank line)
	data = [];
	template = SCHEMA.get_class(target_class);

	with open(temp_base + '.tsv', 'w', newline='') as tsv_file:

		writer = csv.DictWriter(tsv_file, fieldnames = header_row, dialect='excel-tab');
		writer.writeheader();

		# We are in the data rows now, which can be written to new file:
		for row in reader:	

			if type(row) is tuple: # xls/xlsx: for value in row:
				row_data = {k: v for k, v in zip(header_row, row)}

			else: # tsv/csv Dict: e.g. row = {'first name': 'foo', ...}, wher keys have to be converted
				row_data = {k: row[v] for k, v in zip(header_row, row)}
				#print("ROW:",row_data)
					
			# Yaml only gets fields that have values, and as well transformation of some data types.
			data.append(getLinkMLTransform(SCHEMA, template, row_data));
			writer.writerow(row_data);

		YAMLDumper().dump(data, temp_base + '.yaml');	
		return data;


# Could be made more efficient by running vertical loop on columns of a given 
# data type.
def getLinkMLTransform(SCHEMA, template, row_data):
		
	data = {};
	for key, val in row_data.items():
		if val: # Only return dict keys that have values.
			slot = template['attributes'][key];
			output_val = val;
			ranges = [];

			if slot['range']:
				ranges = [slot['range']];
			else:
				for range_type in ['any_of','exactly_one_of','none_of','all_of']:
					if slot[range_type]:
						# e.g. 'any_of': [AnonymousSlotExpression({'range': 'decimal'}),
  					#  AnonymousSlotExpression({'range': 'NullValueMenu'})]
						ranges = [binding.range for binding in slot[range_type]]
						break;

			# ISSUE: If a slot is integer or decimal but value is saved as a
			# string in yaml file, linkml-validate throws error. Must adjust saved
			# datatype
			for slot_range in ranges:
				match slot_range:
					case 'boolean':
						if val.lower() in BOOLEANS:
							output_val = bool(BOOLEANS[val.lower()]);
					case 'integer': 
						if isInteger(val): 
							output_val = int(val);
					case 'decimal'|'float':
						if isDecimal(val): # Note .isdecimal() does NOT test for decimals.
							if '.' in val:
								output_val = float(val); 
							else:
								output_val = int(val);
					#case 'date': 
					case _: # Nothing to do
						pass

			if slot['multivalued'] == True:
				output_val = [x.strip() for x in re.split(DELIMITERS, output_val)];

			# For validation, LinkML will transform both schema and slot labels into
			# what it considers are standardized names, so we have to anticipate what
			# new slot label will be via search and replace.  Convert keys to 
			# **snake_case** since linkml-validate insists on that. However: 
			# - Forward slashes and parentheses are preserved though this is 
			# nonstandard, so:
			# 		"geo_loc name (state/province/territory)"
			# 	is changed to
			#			"geo_loc_name_(state/province/territory)"
			# - Case is preserved though that is non-standard.  So 
			# 	   "specimen collector sample ID" 
			# 	is changed to 
			#      "specimen_collector_sample_ID"
			#
			# - Validating caps CamelCase Enums is hard, e.g. if an Enum is named 
			# 			"geo_loc_name (state/province/territory) menu"
			#		LinkML will automatically rename this to 
			#			"GeoLocName(state/province/territory)Menu"
			#		However, it doesn't update the name in slot range expressions!
			#   Hence these must be renamed in source schema.

			key = re.sub("[-]","",re.sub("[ ]","_", key)); # Accepts ()/ in field name.
			data[key] = output_val;

	return data;

def isDecimal(x):
    try:
        float(x);
        return True
    except ValueError:
        return False

def isInteger(x):
    try: 
        int(x)
    except ValueError:
        return False
    else:
        return True

###############################################################################

warnings = [];
DELIMITERS = '[;|]'; # regex for delimiters in multivalued fields.

args = init_parser();

if not os.path.isfile(args.schema_path):
  exit("LinkML schema file not found: " + args.schema_path)

with open(args.schema_path, "r") as schema_handle:

	# Using SchemaView() to generate inferred slot attributes
	# like in schema.json generated by tabular_to_schema.py
	# Converts schema as javascript object into LinkML schemaView object;
	schema_obj = yaml.safe_load(schema_handle);

	if not "classes" in schema_obj or not "slots" in schema_obj:
		exit("Given schema is missing classes or slots.");

	SCHEMA = SchemaView(yaml.dump(schema_obj, sort_keys=False)); 
	# Brings in any "imports:". This also includes built-in linkml:types
	SCHEMA.merge_imports();

	# Loop through each class and replace it with its induced version which
	# includes attributes dictionary containing inferred slot definitions.
	for name, class_obj in SCHEMA.all_classes().items():
		# Note classDef["@type"]: "ClassDefinition" is only in json output
		if SCHEMA.class_slots(name):
			new_obj = SCHEMA.induced_class(name);
			SCHEMA.add_class(new_obj);
		
	(target_class, slot_key) = getTargetClass(SCHEMA, args.target_class);
	# HAD TO Normalize target class : "CanCoGEN Covid-19" => "CanCOGeNCovid19"

	# Cycle through each data_source file to validate
	for data_source in args.data_sources.split():

		print ("VALIDATING: ", data_source);

		file_path_obj = pathlib.Path(data_source);
		if file_path_obj.suffix in ['.json','.json-ld','.yaml','.yml']:

			# FUTURE: Handle slot name / title variations here too that be 
			# encountered when using a newer or older schema.
			temp_file = data_source;

		else:
			# Deal with section headers and column headers as titles
			# Writes a temporary file with all fields renamed
			temp_file = file_path_obj.stem + ".tmp";
			(yaml, report) = getNormalizedDataFile(SCHEMA, target_class, data_source, temp_file);

			#for item in report:
			#	print (report[item]);

		subprocess.run(["linkml-validate", "-s", args.schema_path, "-C", target_class, temp_file + '.yaml']);			# input='foobar'.encode('utf-8')
		
		print ("File scan complete.");

if len(warnings):
	print ("\nWARNING: \n", "\n ".join(warnings));

"""
# SNIPETS:
# A challenge trying to get linkml-validate working via python module.

report = validate(data_handle, args.schema_path, "Person") # , "Person"
if not report.results:
  print('The instance is valid!')
else:
  for result in report.results:
      print(result.message)

# ISSUE TRYING TO IMPORT linkml, getting "error: no such option: --schema" when --schema provided to argparse; and "Input LinkML schema file not given" when not given that parameter!
try:
	from linkml.validator import validate
except Exception as inst:
	print(type(inst))    # the exception type
	print(inst.args)     # arguments stored in .args
	print(inst)          # __str__ allows args to be printed directly,
	                     # but may be overridden in exception subclasses

	
stderr=None;
try:
# e.g. > linkml-convert -s schema.yaml -C CanCOGeNCovid19 --index-slot specimen_collector_sample_id -o validTestData_2-1-2.tmp.tsv.json validTestData_2-1-2.tmp.tsv
# ISSUE IS range="ANY_OF" slots may have content but if REQUIRED=True, ARE THROWING ERROR.
	run_state = subprocess.check_output(["linkml-convert", "-s", args.schema_path, "-C", target_class, "--index-slot", slot_key, "-o", temp_file + '.json', temp_file]) #, stderr=subprocess.STDOUT

except BaseException as inst:

	print(type(inst))		# the exception type
	print(inst.args)		# arguments stored in .args
	print(inst)					# __str__ allows args to be printed directly,
												# but may be overridden in exception subclasses

finally:
...

"""
